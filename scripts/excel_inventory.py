#!/usr/bin/env python3
"""
Excel Inventory Script (read-only)
---------------------------------
Scans a workspace for Excel workbooks (.xlsx, .xls) and extracts a
sheet-level metadata snapshot without modifying any data.

Outputs:
- analysis/excel_inventory/inventory.json : one entry per sheet with
    file path, sheet name, row/col counts, header info, sample data, etc.
- analysis/excel_inventory/data_dictionary.json : a lightweight canonical
    mapping of raw headers to canonical names (lowercase, underscores).
- analysis/excel_inventory/summary.txt : high-level summary of findings.
"""

from __future__ import annotations

import json
import sys
import os
import re
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

ROOT = Path(__file__).resolve().parents[1]  # repository root
WORKSPACE = Path.cwd()  # assume script is run from workspace root

# Output directory
OUTPUT_ROOT = Path("analysis") / "excel_inventory"

NOTE_FORMAT = {
    "required": 0,
}


def ensure_output_dir():
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)


def safe_json(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False, indent=2, sort_keys=True)


def guess_header_names(first_row_values: List[Any]) -> Optional[List[str]]:
    # Header is considered present if all first row values are non-empty strings
    if not first_row_values:
        return None
    if all(isinstance(v, str) and v is not None and v != "" for v in first_row_values):
        return [str(v).strip() for v in first_row_values]
    return None


def normalize_header(h: str) -> str:
    h = h.strip().lower()
    h = re.sub(r"[^a-z0-9]+", "_", h)
    h = re.sub(r"_+$", "", h)
    h = re.sub(r"^_+", "", h)
    return h if h else "col"


def infer_type_from_value(v: Any) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    if isinstance(v, (datetime,)):
        return "date"
    # Strings may encode numbers or dates; keep as string for now
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return "null"
        # Quick date check
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%Y/%m/%d"):
            try:
                datetime.strptime(s, fmt)
                return "date"
            except Exception:
                pass
        return "string"
    return "string"


def summarize_types(type_list: List[str]) -> str:
    # Return a simple dominant type for a column
    if not type_list:
        return "unknown"
    unique = set(type_list)
    if len(unique) == 1:
        return unique.pop()
    # If mixed numeric types, prefer float, else string
    if unique.issuperset({"int", "float"}):
        return "float"
    return "string"


def process_xlsx(fp: Path) -> Optional[Dict[str, Any]]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(fp.as_posix(), read_only=True, data_only=True)
    except Exception as e:
        # Could be password-protected or unsupported file
        return {
            "file_path": str(fp),
            "workbook_name": fp.name,
            "error": f"cannot_open_xlsx: {e}",
            "password_protected": any(
                ["password" in str(e).lower(), "encrypted" in str(e).lower()]
            ),
            "sheets": [],
        }

    results = []
    for idx, sheet in enumerate(wb.worksheets, start=1):
        sheet_name = sheet.title
        # visibility
        visible = getattr(sheet, "sheet_state", "visible") == "visible"
        max_row = int(sheet.max_row)
        max_col = int(sheet.max_column)
        header_values = []
        try:
            header_values = [
                sheet.cell(row=1, column=c).value for c in range(1, max_col + 1)
            ]
        except Exception:
            header_values = []
        header_names = guess_header_names(header_values) if header_values else None

        header_present = header_names is not None
        if header_present:
            cleaned_header_names = header_names
        else:
            cleaned_header_names = [f"col{c + 1}" for c in range(max_col)]

        # Sample data rows
        sample_rows = []
        data_start = 2 if header_present else 1
        max_samples = 10
        for r in range(data_start, min(data_start + max_samples, max_row + 1)):
            row_vals = []
            row_dict = {}
            for c in range(1, max_col + 1):
                val = None
                try:
                    val = sheet.cell(row=r, column=c).value
                except Exception:
                    val = None
                header_key = (
                    cleaned_header_names[c - 1]
                    if c - 1 < len(cleaned_header_names)
                    else f"col{c}"
                )
                row_dict[header_key] = val
                row_vals.append(val)
            sample_rows.append(row_dict)

        # Infer data types per column from sample
        per_col_types: Dict[str, List[str]] = {
            header: [] for header in cleaned_header_names
        }
        for row in sample_rows:
            for i, header in enumerate(cleaned_header_names):
                val = row.get(header)
                per_col_types[header].append(infer_type_from_value(val))
        data_type_hints = {
            header: summarize_types(types) for header, types in per_col_types.items()
        }

        # Basic anomalies (simple heuristics)
        anomalies: List[str] = []
        if max_row <= 1:
            anomalies.append("empty_or_no_data_rows")
        if header_present and len(set(header_names)) != len(header_names):
            anomalies.append("duplicate_header_values")

        results.append(
            {
                "file_path": str(fp),
                "workbook_name": fp.name,
                "sheet_name": sheet_name,
                "sheet_index": idx,
                "visible": visible,
                "max_row": max_row,
                "max_column": max_col,
                "header_present": header_present,
                "header_names": header_names if header_present else None,
                "data_type_hints": data_type_hints,
                "sample_rows": sample_rows,
                "anomalies": anomalies,
                "password_protected": False,
            }
        )

    wb.close()
    return {
        "file_path": str(fp),
        "workbook_name": fp.name,
        "sheets": results,
    }


def process_xls(fp: Path) -> Optional[Dict[str, Any]]:
    try:
        import xlrd

        rb = xlrd.open_workbook(
            fp.as_posix()
        )  # on_demand not available in xlrd for xls
    except Exception as e:
        return {
            "file_path": str(fp),
            "workbook_name": fp.name,
            "error": f"cannot_open_xls: {e}",
            "password_protected": False,
            "sheets": [],
        }

    results = []
    for idx in range(rb.nsheets):
        sh = rb.sheet_by_index(idx)
        sheet_name = sh.name
        max_row = sh.nrows
        max_col = sh.ncols

        # Header detection
        header_values = sh.row_values(0) if max_row > 0 else []
        header_names = guess_header_names(header_values)
        header_present = header_names is not None
        if header_present:
            cleaned_header_names = header_names
        else:
            cleaned_header_names = [f"col{c + 1}" for c in range(max_col)]

        sample_rows = []
        data_start = 2 if header_present else 1
        max_samples = 10
        for r in range(data_start, min(data_start + max_samples, max_row)):
            row_vals = sh.row_values(r)
            row_dict = {}
            for c in range(max_col):
                val = row_vals[c] if c < len(row_vals) else None
                header_key = (
                    cleaned_header_names[c]
                    if c < len(cleaned_header_names)
                    else f"col{c + 1}"
                )
                row_dict[header_key] = val
            sample_rows.append(row_dict)

        per_col_types = {header: [] for header in cleaned_header_names}
        for row in sample_rows:
            for i, header in enumerate(cleaned_header_names):
                val = row.get(header)
                per_col_types[header].append(infer_type_from_value(val))
        data_type_hints = {
            header: summarize_types(types) for header, types in per_col_types.items()
        }

        anomalies: List[str] = []
        if max_row <= 0:
            anomalies.append("empty_sheet")
        if header_present and len(set(header_names)) != len(header_names):
            anomalies.append("duplicate_header_values")

        results.append(
            {
                "file_path": str(fp),
                "workbook_name": fp.name,
                "sheet_name": sheet_name,
                "sheet_index": idx + 1,
                "visible": True,
                "max_row": max_row,
                "max_column": max_col,
                "header_present": header_present,
                "header_names": header_names if header_present else None,
                "data_type_hints": data_type_hints,
                "sample_rows": sample_rows,
                "anomalies": anomalies,
                "password_protected": False,
            }
        )

    return {
        "file_path": str(fp),
        "workbook_name": fp.name,
        "sheets": results,
    }


def main():
    ensure_output_dir()
    output_inventory = []
    all_sheets_count = 0
    all_workbooks = 0
    errors = []

    root = Path.cwd()
    xlsx_files = sorted(set(p.resolve() for p in root.rglob("*.xlsx")))
    xls_files = sorted(set(p.resolve() for p in root.rglob("*.xls")))

    for fp in xlsx_files:
        all_workbooks += 1
        res = process_xlsx(fp)
        if not res:
            continue
        # accumulate
        for sheet in res.get("sheets", []):
            output_inventory.append(sheet)
        all_sheets_count += len(res.get("sheets", []))

    for fp in xls_files:
        all_workbooks += 1
        res = process_xls(fp)
        if not res:
            continue
        for sheet in res.get("sheets", []):
            output_inventory.append(sheet)
        all_sheets_count += len(res.get("sheets", []))

    # If nothing found, exit gracefully
    if not output_inventory:
        print("No Excel workbooks found in workspace.")
        return

    # Build simple data dictionary from headers
    raw_headers = []
    for item in output_inventory:
        if item.get("header_present") and item.get("header_names"):
            for h in item["header_names"]:
                if isinstance(h, str) and h:
                    raw_headers.append(h)
    canonical_map = {}
    seen = {}
    for h in raw_headers:
        canon = normalize_header(h)
        # resolve duplicates by appending numeric suffix
        if canon in seen:
            seen[canon] += 1
            canon = f"{canon}_{seen[canon]}"
        else:
            seen[canon] = 0
        canonical_map[h] = canon

    data_dictionary = []
    for raw, canon in canonical_map.items():
        data_dictionary.append(
            {"raw_header": raw, "canonical_header": canon, "notes": ""}
        )

    inventory_path = OUTPUT_ROOT / "inventory.json"
    dict_path = OUTPUT_ROOT / "data_dictionary.json"
    summary_path = OUTPUT_ROOT / "summary.txt"

    with open(inventory_path, "w", encoding="utf-8") as f:
        json.dump(output_inventory, f, ensure_ascii=False, indent=2, sort_keys=True)

    with open(dict_path, "w", encoding="utf-8") as f:
        json.dump(data_dictionary, f, ensure_ascii=False, indent=2, sort_keys=True)

    # Simple human-readable summary
    total_sheets = len(output_inventory)
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(f"Excel Inventory Summary\n")
        f.write(f"Generated: {datetime.utcnow().isoformat()}Z\n")
        f.write(f"Workbooks scanned (xlsx/xls): {all_workbooks}\n")
        f.write(f"Total sheets discovered: {total_sheets}\n")
        f.write("\nTop headers observed (sample):\n")
        # sample 5 most frequent header names
        header_counts = {}
        for item in output_inventory:
            if item.get("header_names"):
                for h in item["header_names"]:
                    header_counts[h] = header_counts.get(h, 0) + 1
        top = sorted(header_counts.items(), key=lambda kv: kv[1], reverse=True)[:5]
        for h, c in top:
            f.write(f"- {h}: seen in {c} sheets\n")

    print("Excel inventory completed.")
    print(f"Inventory: {inventory_path}")
    print(f"Dictionary: {dict_path}")
    print(f"Summary: {summary_path}")


if __name__ == "__main__":
    main()
